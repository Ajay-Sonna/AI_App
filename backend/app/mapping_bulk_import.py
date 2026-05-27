"""Parse Excel/CSV bulk column-mapping files and apply upserts (Mapping tab)."""

from __future__ import annotations

import csv
import io
import json
import logging
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Set, Tuple

from app.app_db import fee_column_mappings_repo
from app.app_db.artifacts_repo import get_artifact_by_id, list_artifacts
from app.dst_db.service import validate_table_name

logger = logging.getLogger(__name__)

_CANON_HEADER_ALIASES = {
    "stateschedule": "state_schedule",
    "state_fee_schedule": "state_schedule",
    "state": "state_schedule",
    "artifactid": "artifact_id",
    "artifact_id": "artifact_id",
    "dstschedule": "dst_schedule",
    "dst_fee_schedule": "dst_schedule",
    "dst_table": "dst_schedule",
    "dst": "dst_schedule",
    "statecolumn": "state_column",
    "state_col": "state_column",
    "dstcolumn": "dst_column",
    "dst_col": "dst_column",
    "action": "action",
    "mode": "action",
}


def _norm_ws(s: str) -> str:
    t = unicodedata.normalize("NFKC", (s or "").strip())
    return " ".join(t.split()).lower()


def _header_to_key(h: str) -> Optional[str]:
    k = _norm_ws(h).replace(" ", "")
    if k in _CANON_HEADER_ALIASES:
        return _CANON_HEADER_ALIASES[k]
    lk = _norm_ws(h)
    for alias, canon in _CANON_HEADER_ALIASES.items():
        if alias.replace(" ", "") == k or _norm_ws(alias) == lk:
            return canon
    return None


def _artifact_schedule_title_row(art: Dict[str, Any]) -> Tuple[Optional[int], str]:
    try:
        aid = int(art.get("artifact_id")) if art.get("artifact_id") is not None else None
    except (TypeError, ValueError):
        aid = None
    slabel = str(art.get("source_label") or "").strip()
    lsk = str(art.get("logical_schedule_key") or "").strip().replace("_", " ")
    fn = str(art.get("original_filename") or "").strip()
    if slabel:
        title = slabel
    elif lsk:
        title = lsk
    elif fn:
        base = fn.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        stem = base.rsplit(".", 1)[0].replace("_", " ") if "." in base else base.replace("_", " ")
        title = stem or base or fn
    else:
        title = f"artifact {aid}" if aid is not None else "Fee schedule"
    return aid, title


def _norm_label(s: str) -> str:
    return _norm_ws(s)


def _candidate_labels_for_artifact(art: Dict[str, Any]) -> Set[str]:
    aid, title = _artifact_schedule_title_row(art)
    out: Set[str] = set()
    for raw in (
        title,
        str(art.get("source_label") or "").strip(),
        str(art.get("logical_schedule_key") or "").strip().replace("_", " "),
        str(art.get("logical_schedule_key") or "").strip(),
        str(art.get("original_filename") or "").strip().rsplit("/", 1)[-1].rsplit(".", 1)[0],
    ):
        if raw:
            out.add(_norm_label(raw))
    if aid is not None:
        out.add(_norm_label(f"artifact:{aid}"))
        out.add(_norm_label(str(aid)))
    return {x for x in out if x}


def _parse_table_rows(raw_bytes: bytes) -> Tuple[List[Dict[str, str]], Optional[str]]:
    blob = raw_bytes.lstrip()
    if not blob:
        return [], "empty_file"
    if blob.startswith(b"%PDF"):
        return [], "unsupported_pdf"

    try:
        import zipfile

        if zipfile.is_zipfile(io.BytesIO(blob)):
            return _parse_xlsx(blob)
    except Exception:
        pass

    try:
        text = raw_bytes.decode("utf-8-sig", errors="replace")
    except Exception:
        text = raw_bytes.decode("latin-1", errors="replace")

    try:
        dia = csv.Sniffer().sniff(text[:8000], delimiters=",\t;")
    except csv.Error:
        dia = csv.excel
    rdr = csv.DictReader(io.StringIO(text), dialect=dia)
    if not rdr.fieldnames:
        return [], "no_csv_header"
    hmap = {_header_to_key(h or ""): h for h in rdr.fieldnames if h}
    hmap = {k: v for k, v in hmap.items() if k}
    if "state_schedule" not in hmap and "artifact_id" not in hmap:
        return [], "missing_state_column"
    if "dst_schedule" not in hmap:
        return [], "missing_dst_schedule_column"
    if "state_column" not in hmap or "dst_column" not in hmap:
        return [], "missing_mapping_columns"

    rows: List[Dict[str, str]] = []
    for line in rdr:
        if not line:
            continue
        rec: Dict[str, str] = {}
        for ck, src_h in hmap.items():
            v = line.get(src_h)
            rec[ck] = str(v).strip() if v is not None else ""
        if any(rec.values()):
            rows.append(rec)
    return rows, None


def _parse_xlsx(raw_bytes: bytes) -> Tuple[List[Dict[str, str]], Optional[str]]:
    try:
        import openpyxl
    except ImportError:
        return [], "openpyxl_missing"
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        grid: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            grid.append([("" if c is None else str(c)).strip() for c in row])
            if len(grid) >= 5000:
                break
        if not grid:
            return [], "empty_sheet"
        header_cells = grid[0]
        hmap: Dict[str, int] = {}
        for idx, cell in enumerate(header_cells):
            ck = _header_to_key(str(cell))
            if ck:
                hmap[ck] = idx
        if "state_schedule" not in hmap and "artifact_id" not in hmap:
            return [], "missing_state_column"
        if "dst_schedule" not in hmap:
            return [], "missing_dst_schedule_column"
        if "state_column" not in hmap or "dst_column" not in hmap:
            return [], "missing_mapping_columns"

        rows: List[Dict[str, str]] = []
        for r in grid[1:]:
            rec: Dict[str, str] = {}
            for ck, i in hmap.items():
                rec[ck] = str(r[i]).strip() if i < len(r) else ""
            if any(rec.values()):
                rows.append(rec)
        return rows, None
    finally:
        wb.close()


@dataclass
class _Group:
    state_schedule: str = ""
    artifact_id: Optional[int] = None
    dst_fsname: str = ""
    action: str = "replace"
    pairs: Dict[str, str] = field(default_factory=dict)
    source_rows: List[int] = field(default_factory=list)


def _build_label_index(artifacts: List[Dict[str, Any]]) -> Dict[str, List[int]]:
    idx: Dict[str, List[int]] = defaultdict(list)
    for art in artifacts:
        aid = art.get("artifact_id")
        if aid is None:
            continue
        try:
            n = int(aid)
        except (TypeError, ValueError):
            continue
        for lab in _candidate_labels_for_artifact(art):
            if lab:
                idx[lab].append(n)
    return idx


def run_bulk_mapping_import(
    *,
    state_code: str,
    raw_bytes: bytes,
    dry_run: bool = False,
    updated_by: Optional[str] = None,
) -> Dict[str, Any]:
    sc = str(state_code or "").strip().upper()[:8]
    if not sc:
        return {
            "ok": False,
            "error": "state_code required",
            "state_code": "",
            "dry_run": dry_run,
            "applied": [],
            "errors": [],
            "warnings": [],
            "groups_total": 0,
        }

    rows, err = _parse_table_rows(raw_bytes)
    if err:
        return {
            "ok": False,
            "error": f"parse_failed:{err}",
            "state_code": sc,
            "dry_run": dry_run,
            "applied": [],
            "errors": [],
            "warnings": [],
            "groups_total": 0,
        }
    if not rows:
        return {
            "ok": False,
            "error": "no_data_rows",
            "state_code": sc,
            "dry_run": dry_run,
            "applied": [],
            "errors": [],
            "warnings": [],
            "groups_total": 0,
        }

    groups: Dict[Tuple[str, str], _Group] = {}
    row_num = 1
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

    for line in rows:
        row_num += 1
        aid_raw = (line.get("artifact_id") or "").strip()
        ss = (line.get("state_schedule") or "").strip()
        dst = (line.get("dst_schedule") or "").strip()
        stc = (line.get("state_column") or "").strip()
        dtc = (line.get("dst_column") or "").strip()
        act = (line.get("action") or "replace").strip().lower() or "replace"

        if not dst or not stc or not dtc:
            errors.append({"row": row_num, "message": "DstSchedule, StateColumn, and DstColumn are required."})
            continue

        aid: Optional[int] = None
        if aid_raw:
            try:
                aid = int(float(aid_raw))
            except (TypeError, ValueError):
                errors.append({"row": row_num, "message": f"Invalid ArtifactId: {aid_raw!r}"})
                continue
            gkey = (f"aid:{aid}", dst)
        else:
            if not ss:
                errors.append({"row": row_num, "message": "StateSchedule or ArtifactId is required."})
                continue
            gkey = (_norm_label(ss), dst)

        g = groups.get(gkey)
        if g is None:
            g = _Group(
                state_schedule=ss,
                artifact_id=aid,
                dst_fsname=dst,
            )
            groups[gkey] = g
        else:
            if aid is not None and g.artifact_id is not None and g.artifact_id != aid:
                errors.append(
                    {
                        "row": row_num,
                        "message": f"Inconsistent ArtifactId for same schedule/DST group (was {g.artifact_id}, now {aid}).",
                    }
                )
                continue
            if aid is not None:
                g.artifact_id = aid
            if ss and not g.state_schedule:
                g.state_schedule = ss

        g.action = "merge" if act.startswith("merg") else "replace"

        prev_dst = g.pairs.get(stc)
        if prev_dst is not None and prev_dst.strip() != dtc:
            warnings.append(
                {
                    "row": row_num,
                    "message": f"Duplicate StateColumn {stc!r}: overwriting {prev_dst!r} → {dtc!r}",
                }
            )
        g.pairs[stc] = dtc
        g.source_rows.append(row_num)

    if errors and not groups:
        return {
            "ok": False,
            "state_code": sc,
            "dry_run": dry_run,
            "errors": errors,
            "warnings": warnings,
            "applied": [],
            "groups_total": 0,
        }

    arts = list_artifacts(state_code=sc, current_only=False, limit=5000)
    label_index = _build_label_index(arts)

    applied: List[Dict[str, Any]] = []

    for gkey, g in groups.items():
        try:
            dst_clean = validate_table_name(g.dst_fsname)
        except ValueError as ve:
            errors.append(
                {
                    "row": min(g.source_rows) if g.source_rows else 0,
                    "message": str(ve),
                }
            )
            continue

        aid = g.artifact_id
        if aid is None:
            lab = str(gkey[0])
            matches = label_index.get(lab, [])
            uniq = sorted(set(matches))
            if not uniq:
                errors.append(
                    {
                        "row": min(g.source_rows) if g.source_rows else 0,
                        "message": f"No saved fee file matches StateSchedule {g.state_schedule!r} for {sc}.",
                    }
                )
                continue
            if len(uniq) > 1:
                errors.append(
                    {
                        "row": min(g.source_rows) if g.source_rows else 0,
                        "message": (
                            f"Multiple artifacts match {g.state_schedule!r}: ids {uniq[:12]}"
                            f"{'…' if len(uniq) > 12 else ''}. Set ArtifactId column to disambiguate."
                        ),
                    }
                )
                continue
            aid = uniq[0]

        row_art = get_artifact_by_id(int(aid))
        if not row_art:
            errors.append(
                {
                    "row": min(g.source_rows) if g.source_rows else 0,
                    "message": f"Artifact {aid} not found.",
                }
            )
            continue
        art_sc = str(row_art.get("state_code") or "").strip().upper()
        if art_sc and art_sc != sc:
            errors.append(
                {
                    "row": min(g.source_rows) if g.source_rows else 0,
                    "message": f"Artifact {aid} belongs to state {art_sc}, not {sc}.",
                }
            )
            continue

        try:
            lsk = fee_column_mappings_repo.resolve_schedule_key_for_artifact(row_art)
        except ValueError as ve:
            errors.append({"row": min(g.source_rows) if g.source_rows else 0, "message": str(ve)})
            continue

        column_map: Dict[str, str] = dict(g.pairs)
        if g.action == "merge":
            exist = fee_column_mappings_repo.lookup_latest_mapping(
                state_code=sc,
                state_logical_schedule_key=lsk,
                dst_fsname=dst_clean,
            )
            if exist:
                prev: Dict[str, str] = {}
                raw_cm = exist.get("column_map_json")
                if isinstance(raw_cm, dict):
                    prev = {str(k): str(v) for k, v in raw_cm.items()}
                elif isinstance(raw_cm, str) and raw_cm.strip():
                    try:
                        j = json.loads(raw_cm)
                        if isinstance(j, dict):
                            prev = {str(k): str(v) for k, v in j.items()}
                    except json.JSONDecodeError:
                        pass
                prev.update(column_map)
                column_map = {k: v for k, v in prev.items() if str(k).strip() and str(v).strip()}

        if dry_run:
            applied.append(
                {
                    "artifact_id": int(aid),
                    "dst_fsname": dst_clean,
                    "logical_schedule_key": lsk,
                    "pairs": len(column_map),
                    "action": g.action,
                    "dry_run": True,
                }
            )
            continue

        try:
            saved = fee_column_mappings_repo.upsert_fee_column_mapping(
                state_code=sc,
                state_logical_schedule_key=lsk,
                dst_fsname=dst_clean,
                column_map_json=column_map,
                updated_by=updated_by,
            )
            applied.append(
                {
                    "artifact_id": int(aid),
                    "mapping_id": int(saved.get("mapping_id")),
                    "dst_fsname": dst_clean,
                    "logical_schedule_key": lsk,
                    "pairs": len(column_map),
                    "action": g.action,
                }
            )
        except Exception as ex:
            logger.exception("bulk mapping upsert failed")
            errors.append(
                {
                    "row": min(g.source_rows) if g.source_rows else 0,
                    "message": str(ex),
                }
            )

    success = len(applied) > 0 and len(errors) == 0
    return {
        "ok": success,
        "state_code": sc,
        "dry_run": dry_run,
        "applied": applied,
        "errors": errors,
        "warnings": warnings,
        "groups_total": len(groups),
    }
