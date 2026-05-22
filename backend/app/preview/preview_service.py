"""
Generic URL preview: SSRF defenses + cookie-authenticated replay + format sniffers.

Nothing is keyed off a specific US state — rules use URL parsing, referrer, and cookie
scopes captured during the originating run.
"""

from __future__ import annotations

import base64
import csv
import ipaddress
import io
import re
import zipfile
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import requests

from app.preview.session_store import StoredPreviewSession

_DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

MAX_FETCH_BYTES = 15 * 1024 * 1024
MAX_INLINE_BASE64_RAW = 4_500_000
MAX_PREVIEW_ROWS = 80
# Saved fee artifacts: return as much of the primary grid as practical (comparison prep).
ARTIFACT_TABLE_MAX_ROWS = 100_000
MAX_TEXT_CHARS = 12_000
_MAX_SCAN_ROWS = 520
_RICH_XLSX_BYTES_CAP = 12 * 1024 * 1024

# Newer first: Akamai / similar stacks often expect current Chrome TLS + UA alignment.
_CURL_IMPERSONATE_CHAIN = (
    "chrome146",
    "chrome136",
    "chrome133a",
    "chrome131",
    "safari184",
)

_RE_MONEY_LIKE = re.compile(r"^\s*\$?\s*[\d,]+\.\d{1,4}\s*$")
_RE_PROC_CODE_LIKE = re.compile(r"^\d{4,8}$")
_RE_SYNTH_COL_HEADER = re.compile(r"^col_\d+$", re.I)


def _dedupe_accept_attempts(
    steps: List[Optional[Dict[str, str]]],
) -> List[Optional[Dict[str, str]]]:
    seen: set[Tuple[str, ...]] = set()
    out: List[Optional[Dict[str, str]]] = []
    for h in steps:
        key = tuple(sorted(h.items())) if h else tuple()
        if key in seen:
            continue
        seen.add(key)
        out.append(h)
    return out if out else [None]


def pdf_preview_accept_attempts(resource_url: str) -> List[Optional[Dict[str, str]]]:
    """
    PDF preview only: neutral session Accept first, then PDF / octet — no Excel MIME here.
    """
    ext = _extra_accept_headers(resource_url)
    steps: List[Optional[Dict[str, str]]] = [
        None,
        {
            "Accept": (
                "application/pdf,application/x-pdf,"
                "application/octet-stream;q=0.95,*/*;q=0.05"
            )
        },
    ]
    if ext:
        steps.append(ext)
    steps.extend(
        [
            {"Accept": "*/*"},
            {"Accept": "application/octet-stream,*/*;q=0.1"},
        ]
    )
    return _dedupe_accept_attempts(steps)


def spreadsheet_preview_accept_attempts(resource_url: str) -> List[Optional[Dict[str, str]]]:
    """Excel / spreadsheet preview only — never sends application/pdf-first."""
    ext = _extra_accept_headers(resource_url)
    steps: List[Optional[Dict[str, str]]] = [
        None,
        {
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/vnd.ms-excel,"
                "application/octet-stream;q=0.95,*/*;q=0.05"
            )
        },
    ]
    if ext:
        steps.append(ext)
    steps.extend(
        [
            {"Accept": "*/*"},
            {"Accept": "application/octet-stream,*/*;q=0.1"},
        ]
    )
    return _dedupe_accept_attempts(steps)


def csv_preview_accept_attempts(resource_url: str) -> List[Optional[Dict[str, str]]]:
    ext = _extra_accept_headers(resource_url)
    steps: List[Optional[Dict[str, str]]] = [
        None,
        {"Accept": "text/csv,text/plain,application/octet-stream;q=0.9,*/*;q=0.05"},
    ]
    if ext:
        steps.append(ext)
    steps.append({"Accept": "*/*"})
    return _dedupe_accept_attempts(steps)


def auto_preview_accept_attempts(resource_url: str) -> List[Optional[Dict[str, str]]]:
    """Unknown type: merge PDF + spreadsheet attempt lists (deduped)."""
    return _dedupe_accept_attempts(
        pdf_preview_accept_attempts(resource_url) + spreadsheet_preview_accept_attempts(resource_url)
    )


def accept_attempts_for_preview(
    resource_url: str,
    document_hint: Optional[str] = None,
) -> List[Optional[Dict[str, str]]]:
    dh = (document_hint or "").strip().lower()
    if dh in ("spreadsheet", "excel", "xlsx", "xls"):
        return spreadsheet_preview_accept_attempts(resource_url)
    if dh == "pdf":
        return pdf_preview_accept_attempts(resource_url)
    if dh == "csv":
        return csv_preview_accept_attempts(resource_url)
    return auto_preview_accept_attempts(resource_url)


def _accept_label(h: Optional[Dict[str, str]]) -> str:
    if not h:
        return "default_session_accept"
    return (h.get("Accept") or "?")[:220]


def _suffix_from_url(resource_url: str) -> str:
    try:
        leaf = (urlparse(resource_url).path or "").rsplit("/", 1)[-1].lower()
    except Exception:
        return ""
    if "." not in leaf:
        return ""
    return "." + leaf.rsplit(".", 1)[-1]


def _extra_accept_headers(resource_url: str) -> Dict[str, str]:
    """Some hosts reject generic */* downloads; advertise the artifact type."""
    suf = _suffix_from_url(resource_url)
    if suf == ".pdf":
        return {
            "Accept": (
                "application/pdf,application/x-pdf,application/octet-stream;q=0.9,*/*;q=0.4"
            )
        }
    if suf in {".xlsx", ".xls", ".xlsm"}:
        return {
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/vnd.ms-excel,application/octet-stream;q=0.9,*/*;q=0.4"
            )
        }
    if suf == ".csv":
        return {"Accept": "text/csv,text/plain,application/octet-stream;q=0.9,*/*;q=0.4"}
    return {}


_PRIVATE_HOST_PATTERNS = (
    r"^localhost$",
    r"^127\.",
    r"^169\.254\.",
    r"^192\.168\.",
    r"^10\.",
    r"^172\.(1[6-9]|2[0-9]|3[0-9])\.",  # 172.16 - 172.31
)


def _host_is_blocked(hostname: str) -> bool:
    h = (hostname or "").strip().lower()
    if not h:
        return True
    for pat in _PRIVATE_HOST_PATTERNS:
        if re.match(pat, h):
            return True
    try:
        ipaddress.ip_address(h)
        if ipaddress.ip_address(h).is_private or ipaddress.ip_address(h).is_loopback:
            return True
    except ValueError:
        pass
    if h in {"metadata.google.internal", "metadata", "kubernetes.default.svc"}:
        return True
    return False


def _cookie_domain_scopes(cookies: List[Dict[str, Any]]) -> List[str]:
    out: List[str] = []
    for c in cookies or []:
        d = (c.get("domain") or "").strip().lower().lstrip(".")
        if d and d not in out:
            out.append(d)
    return out


def _hosts_share_trust_scope(resource_host: str, referrer_host: str, cookies: List[Dict[str, Any]]) -> bool:
    """Require resource + referrer hosts to overlap an observed cookie-domain scope."""
    rh = resource_host.strip().lower()
    ref = referrer_host.strip().lower()
    if not rh or not ref:
        return rh == ref
    if rh == ref:
        return True
    for scope in _cookie_domain_scopes(cookies):
        r_ok = ref == scope or ref.endswith("." + scope)
        h_ok = rh == scope or rh.endswith("." + scope)
        if r_ok and h_ok:
            return True
    return False


def validate_preview_url(resource_url: str, authority: StoredPreviewSession | None) -> Tuple[bool, Optional[str]]:
    try:
        p = urlparse(resource_url.strip())
    except Exception:
        return False, "invalid_url"

    if p.scheme not in ("http", "https"):
        return False, "unsupported_scheme"

    host = (p.hostname or "").lower()
    if not host or _host_is_blocked(host):
        return False, "blocked_host"

    ul = resource_url.strip().lower()

    if authority:
        refh = (urlparse(authority.referrer_url).hostname or "").lower()
        if not refh or _host_is_blocked(refh):
            return False, "bad_referrer"
        if host == refh:
            pass
        elif _cookie_domain_scopes(authority.cookies) and _hosts_share_trust_scope(host, refh, authority.cookies):
            pass
        else:
            return False, "host_not_in_preview_scope"
    else:
        if _looks_like_sensitive_internal_url(ul):
            return False, "credential_fetch_requires_preview_session"

    return True, None


def _looks_like_sensitive_internal_url(ul_raw: str) -> bool:
    ul = ul_raw.strip().lower()
    if "pubatt/dl/" in ul or "/sys_attachment.do" in ul:
        return True
    if "/api/" in ul and "/now/" in ul:
        return True
    return False


def session_from_browser(
    authority: StoredPreviewSession | None,
    *,
    extra_referrer: Optional[str] = None,
) -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": _DEFAULT_UA,
            "Accept": "*/*",
        }
    )
    csrf_token: Optional[str] = None
    cookies_src: List[Dict[str, Any]] = []
    scoped_ref_url = ""

    if authority:
        cookies_src = authority.cookies or []
        csrf_token = authority.csrf_token
        scoped_ref_url = authority.referrer_url or ""

    referrer = (extra_referrer or "").strip() or scoped_ref_url
    if referrer:
        s.headers["Referer"] = referrer
        # Do NOT set Origin. Many static file / media CDNs (incl. some .gov portals) reject
        # downloads that include a browser Origin header on simple GET requests.

    if csrf_token:
        s.headers["X-UserToken"] = csrf_token

    for c in cookies_src:
        name = c.get("name")
        if not name:
            continue
        value = c.get("value") or ""
        domain = c.get("domain") or ""
        path = c.get("path") or "/"
        try:
            s.cookies.set(name, value, domain=domain, path=path)
        except Exception:
            s.cookies.set(name, value, path=path)

    return s


def _xml_body_indicates_auth_failure(data: bytes) -> bool:
    head = data[:4096].lower()
    if not head.strip().startswith(b"<"):
        return False
    compact = head.replace(b"\n", b"").replace(b" ", b"")
    if b"usernotauthenticated" in compact or b"notauthenticated" in compact:
        return True
    if b"<status>failure</status>" in compact and b"<error>" in compact:
        return True
    return False


def _registrable_e2ld(host: str) -> str:
    """Naïve effective TLD+1 for Sec-Fetch-Site (good enough for .gov / most US state sites)."""
    h = (host or "").lower().split(":")[0].strip(".")
    parts = [p for p in h.split(".") if p]
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return h


def _sec_fetch_site_value(resource_url: str, referer: str) -> str:
    ru, rf = urlparse(resource_url), urlparse(referer or "")
    rh, fh = (ru.hostname or "").lower(), (rf.hostname or "").lower()
    if not rh or not fh:
        return "none"
    if rh == fh and (ru.port or 0) == (rf.port or 0) and (ru.scheme or "") == (rf.scheme or ""):
        return "same-origin"
    if _registrable_e2ld(rh) == _registrable_e2ld(fh):
        return "same-site"
    return "cross-site"


def _browser_client_headers_for_media_get(resource_url: str, referer: str) -> Dict[str, str]:
    """Headers browsers send on file GETs; paired with Referer helps some WAFs."""
    site = _sec_fetch_site_value(resource_url, referer)
    return {
        "Accept-Language": "en-US,en;q=0.9",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": site,
        "Sec-Fetch-User": "?1",
    }


def _merge_accept_with_browser_headers(
    resource_url: str,
    referer: str,
    accept_headers: Optional[Dict[str, str]],
) -> Dict[str, str]:
    base = _browser_client_headers_for_media_get(resource_url, referer)
    if accept_headers:
        base.update(accept_headers)
    return base


def _session_cookies_simple_dict(session: requests.Session) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for c in session.cookies:
        try:
            out[c.name] = c.value
        except Exception:
            continue
    return out


def _maybe_warm_referrer(
    session: requests.Session,
    resource_url: str,
    referer: str,
    attempt_log: Optional[List[Dict[str, Any]]],
) -> None:
    ref = (referer or "").strip()
    if not ref.lower().startswith("http"):
        return
    ru, rf = urlparse(resource_url), urlparse(ref)
    rh, fh = (ru.hostname or "").lower(), (rf.hostname or "").lower()
    # Same host or same registrable site (subdomains / CDN vs portal host)
    if rh != fh and _registrable_e2ld(rh) != _registrable_e2ld(fh):
        return
    try:
        r = session.get(ref, timeout=25, allow_redirects=True)
        if attempt_log is not None:
            attempt_log.append(
                {
                    "accept_strategy": "warm_referrer_html",
                    "http_status": r.status_code,
                    "content_type": (r.headers.get("content-type") or "")[:120],
                }
            )
    except Exception as ex:
        if attempt_log is not None:
            attempt_log.append(
                {
                    "accept_strategy": "warm_referrer_html",
                    "http_status": -1,
                    "content_type": str(ex)[:160],
                }
            )


def _should_warm_referrer_page(resource_url: str, referer: str) -> bool:
    ref = (referer or "").strip()
    if not ref.lower().startswith("http"):
        return False
    ru, rf = urlparse(resource_url), urlparse(ref)
    rh, fh = (ru.hostname or "").lower(), (rf.hostname or "").lower()
    return rh == fh or _registrable_e2ld(rh) == _registrable_e2ld(fh)


def _curl_warm_referrer_merge_cookies(
    resource_url: str,
    referer: str,
    cookie_dict: Dict[str, str],
    attempt_log: List[Dict[str, Any]],
) -> Dict[str, str]:
    """
    When ``requests`` warms the listing page with 403, WAF cookies never land in the jar.
    Re-fetch the referrer HTML with curl TLS impersonation and merge Set-Cookie into ``cookie_dict``.
    """
    if not _should_warm_referrer_page(resource_url, referer):
        return cookie_dict
    ref = (referer or "").strip()
    try:
        from curl_cffi import requests as creq
    except ImportError:
        return cookie_dict

    ck = dict(cookie_dict)
    html_accept = {
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,image/apng,*/*;q=0.8"
        )
    }
    hdrs = _merge_accept_with_browser_headers(ref, referer, html_accept)
    for imp in _CURL_IMPERSONATE_CHAIN:
        label = f"warm_referrer_curl_{imp}"
        try:
            r = creq.get(
                ref,
                headers=hdrs,
                cookies=ck or None,
                timeout=60,
                allow_redirects=True,
                impersonate=imp,
            )
        except Exception as ex:
            attempt_log.append(
                {
                    "accept_strategy": label,
                    "http_status": -1,
                    "content_type": str(ex)[:120],
                }
            )
            continue
        ct = (r.headers.get("content-type") or "").split(";")[0].strip() or ""
        attempt_log.append(
            {
                "accept_strategy": label,
                "http_status": r.status_code,
                "content_type": ct[:120],
            }
        )
        try:
            for name, value in r.cookies.items():
                ck[name] = value
        except Exception:
            pass
        if r.status_code < 400:
            break
    return ck


def _download_via_curl_cffi(
    resource_url: str,
    referer: str,
    cookie_dict: Dict[str, str],
    accept_headers: Optional[Dict[str, str]],
    attempt_log: List[Dict[str, Any]],
    *,
    max_response_bytes: Optional[int] = None,
) -> Optional[Tuple[bytes, str, int]]:
    """TLS impersonation fallback when plain ``requests`` gets 403 + HTML."""
    cap = max_response_bytes if max_response_bytes is not None else MAX_FETCH_BYTES
    try:
        from curl_cffi import requests as creq
    except ImportError:
        attempt_log.append(
            {
                "accept_strategy": "curl_cffi_import_error",
                "http_status": -1,
                "content_type": "pip install curl-cffi",
            }
        )
        return None

    for imp in _CURL_IMPERSONATE_CHAIN:
        hdrs = _merge_accept_with_browser_headers(resource_url, referer, accept_headers)
        label = f"curl_cffi_{imp}"
        try:
            r = creq.get(
                resource_url,
                headers=hdrs,
                cookies=cookie_dict or None,
                timeout=90,
                allow_redirects=True,
                impersonate=imp,
            )
        except Exception as ex:
            attempt_log.append(
                {
                    "accept_strategy": label,
                    "http_status": -1,
                    "content_type": str(ex)[:120],
                }
            )
            continue

        ct = (r.headers.get("content-type") or "").split(";")[0].strip() or ""
        attempt_log.append(
            {
                "accept_strategy": label,
                "http_status": r.status_code,
                "content_type": ct[:120],
            }
        )
        if r.status_code >= 400:
            continue
        data = r.content
        if len(data) > cap:
            continue
        return data, ct, r.status_code

    return None


def fetch_resource_bytes(
    resource_url: str,
    authority: StoredPreviewSession | None,
    *,
    referer_override: Optional[str] = None,
    document_hint: Optional[str] = None,
) -> Tuple[bytes, Optional[str], Optional[int], Optional[str], List[Dict[str, Any]]]:
    referer = (referer_override or "").strip()
    if not referer and authority:
        referer = (authority.referrer_url or "").strip()

    session = session_from_browser(authority, extra_referrer=referer_override)
    attempts = accept_attempts_for_preview(resource_url, document_hint=document_hint)
    attempt_log: List[Dict[str, Any]] = []
    last_status: Optional[int] = None
    last_ctype: Optional[str] = None

    _maybe_warm_referrer(session, resource_url, referer, attempt_log)

    for hdrs in attempts:
        merged = _merge_accept_with_browser_headers(resource_url, referer, hdrs)
        resp = session.get(resource_url, stream=True, timeout=60, headers=merged)
        status = resp.status_code
        ctype = resp.headers.get("content-type", "").split(";")[0].strip() or None
        ctype_short = (ctype or "")[:120]
        attempt_log.append(
            {
                "accept_strategy": _accept_label(hdrs),
                "http_status": status,
                "content_type": ctype_short,
            }
        )
        if status >= 400:
            resp.close()
            last_status = status
            last_ctype = ctype
            continue

        chunks: List[bytes] = []
        total = 0
        try:
            for chunk in resp.iter_content(chunk_size=65_536):
                if chunk:
                    total += len(chunk)
                    if total > MAX_FETCH_BYTES:
                        resp.close()
                        return b"", None, None, None, attempt_log
                    chunks.append(chunk)
        except Exception:
            resp.close()
            return b"", ctype, None, None, attempt_log

        combined = b"".join(chunks)
        eff = status
        upstream_hint: Optional[str] = None
        if status < 400 and combined and _xml_body_indicates_auth_failure(combined):
            eff = 401
            upstream_hint = "auth_xml"
        return combined, ctype, eff, upstream_hint, []

    # Plain requests + browser headers still blocked (common: Akamai / bot 403 + HTML)
    ck = _session_cookies_simple_dict(session)
    ck = _curl_warm_referrer_merge_cookies(resource_url, referer, ck, attempt_log)
    dh = (document_hint or "").strip().lower()
    curl_tries: List[Optional[Dict[str, str]]] = [
        None,
        {
            "Accept": (
                "application/pdf,application/x-pdf,"
                "application/octet-stream;q=0.95,*/*;q=0.05"
            )
        },
    ]
    if dh in ("spreadsheet", "excel", "xlsx", "xls"):
        curl_tries.append(
            {
                "Accept": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                    "application/vnd.ms-excel,"
                    "application/octet-stream;q=0.95,*/*;q=0.05"
                )
            }
        )

    for caccept in curl_tries:
        hit = _download_via_curl_cffi(resource_url, referer, ck, caccept, attempt_log)
        if not hit:
            continue
        combined, ctype, _st = hit
        if combined and _xml_body_indicates_auth_failure(combined):
            return b"", ctype, 401, "auth_xml", attempt_log
        return combined, ctype, _st, None, []

    return b"", last_ctype, last_status or 502, None, attempt_log


def _is_ooxlsx_workbook(data: bytes) -> bool:
    try:
        if not zipfile.is_zipfile(io.BytesIO(data)):
            return False
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = frozenset(zf.namelist())
        return "[Content_Types].xml" in names and (
            "xl/workbook.xml" in names or "xl/workbook.bin" in names
        )
    except Exception:
        return False


def _kind_from_signals(
    content_type: Optional[str],
    url: str,
    data: bytes,
) -> str:
    ct = (content_type or "").lower()
    pu = urlparse(url)
    path = (pu.path or "").lower()
    sniff = data[:16]

    if sniff.startswith(b"%PDF") or ct == "application/pdf":
        return "pdf"

    if (
        "spreadsheet" in ct
        or "excel" in ct
        or path.endswith(".xlsx")
        or path.endswith(".xls")
        or _is_ooxlsx_workbook(data)
    ):
        return "spreadsheet"

    if "csv" in ct or path.endswith(".csv"):
        return "tabular_text"

    if ct.startswith("text/") or path.endswith(".txt"):
        return "text"

    if sniff.startswith((b"<html", b"<!DOC", b"{")):
        return "probably_html_json"

    return "binary"


def _nonempty_width(row: List[Any]) -> int:
    return sum(1 for c in row if str(c).strip())


def _ws_scan_last_row(ws: Any, scan_cap: int) -> int:
    """Prefer worksheet dimension over max_row alone — some NY schedules report max_row too low."""
    mr = int(ws.max_row or 0)
    try:
        from openpyxl.utils import range_boundaries

        dim = ws.calculate_dimension()
        if dim:
            _min_c, _min_r, _max_c, mr_dim = range_boundaries(dim)
            mr = max(mr, int(mr_dim or 0))
    except Exception:
        pass
    return max(2, min(max(mr, 2), scan_cap))


def _looks_like_pure_number_text(t: str) -> bool:
    s = str(t).strip().replace(",", "")
    if not s or s in ("-", "."):
        return False
    try:
        float(s)
        return "e" not in s.lower()
    except ValueError:
        return False


def _cell_looks_like_rate_identifier(t: str) -> bool:
    """NY Medicaid-style codes such as 14110HR, 19340NYC."""
    t = str(t).strip()
    if not t:
        return False
    return bool(re.match(r"^\d{4,6}[A-Z][A-Za-z0-9-]*$", t))


def _row_numeric_like_count(row: List[Any]) -> int:
    return sum(1 for c in row if _looks_like_pure_number_text(str(c)))


def _row_rate_identifier_count(row: List[Any]) -> int:
    return sum(1 for c in row if _cell_looks_like_rate_identifier(str(c)))


def _below_header_data_signal(rows: List[List[Any]], start: int, window: int = 14) -> Tuple[int, int]:
    """
    Strength of tabular-looking rows shortly after a candidate header.
    Handles blank spacer rows between the header band and the first data row.
    """
    if start + 1 >= len(rows):
        return 0, 0
    end = min(start + 1 + window, len(rows))
    scores = [_row_data_likeness(rows[j]) for j in range(start + 1, end)]
    if not scores:
        return 0, 0
    top = sorted(scores, reverse=True)
    pair = top[0] + (top[1] if len(top) > 1 else 0)
    return pair, top[0]


def _cell_looks_like_tabular_data(s: str) -> bool:
    t = str(s).strip()
    if not t:
        return False
    if _RE_MONEY_LIKE.match(t):
        return True
    ts = t.replace(",", "")
    if re.match(r"^-?\d+\.\d{2,}$", ts):
        return True
    if _cell_looks_like_rate_identifier(t):
        return True
    if _RE_PROC_CODE_LIKE.match(t):
        return True
    if 1 <= len(t) <= 4 and re.match(r"^[A-Za-z]{1,4}$", t):
        return True
    if re.match(r"^\d{1,3}$", t):
        return True
    return False


def _row_data_likeness(row: List[Any]) -> int:
    return sum(1 for c in row if _cell_looks_like_tabular_data(str(c)))


def _header_keyword_bonus(row: List[Any]) -> float:
    """Boost detection of fee-schedule grids (multi-row merged headers confuse generic scoring)."""
    s = " ".join(str(c).strip().lower() for c in row if c is not None and str(c).strip())
    if not s:
        return 0.0
    bonus = 0.0
    if re.search(r"\bprogram\b", s):
        bonus += 22.0
    if "workforce" in s or "recruitment" in s or "retention" in s:
        bonus += 14.0
    if "hcbs" in s or "resc" in s or "american rescue" in s:
        bonus += 14.0
    if ("rate" in s and "inc" in s) or "effective" in s or "eff. date" in s:
        bonus += 10.0
    if "fee" in s and "schedule" in s:
        bonus += 6.0
    if re.search(r"\brate\s+code\b", s) or ("downstate" in s and "count" in s):
        bonus += 14.0
    if "regional rate" in s or "provider-specific" in s or "community residence" in s:
        bonus += 8.0
    if re.search(r"\bcount(y|ies)\b", s) and len(s) > 12:
        bonus += 6.0
    return bonus


def _combine_two_header_rows(prev: List[Any], cur: List[Any], width: int) -> List[str]:
    a = _normalize_row_len(prev, width)
    b = _normalize_row_len(cur, width)
    out: List[str] = []
    for i in range(width):
        p = str(a[i]).strip()
        q = str(b[i]).strip()
        numeric_pair = _looks_like_pure_number_text(p) and _looks_like_pure_number_text(q)
        id_pair = _cell_looks_like_rate_identifier(p) and _cell_looks_like_rate_identifier(q)
        if numeric_pair or id_pair:
            # Avoid "amount · amount" / "code · code" from mistakenly merged data rows.
            out.append(q or p)
            continue
        if p and q and p.lower() != q.lower():
            out.append(f"{p} · {q}")
        else:
            out.append(p or q)
    return out


def _header_pair_merge_worthy(prev: List[Any], cur: List[Any], width: int, grid: List[List[Any]], idx: int) -> bool:
    """True when two consecutive rows are a spreadsheet title/header band above data."""
    if idx <= 0 or idx >= len(grid) - 1:
        return False
    p = _normalize_row_len(prev, width)
    q = _normalize_row_len(cur, width)
    if _nonempty_width(p) < 2 or _nonempty_width(q) < 3:
        return False
    # Prev row must not resemble a provider/rate grid row (two-row merge heuristics).
    if _row_rate_identifier_count(p) >= 2 or _row_numeric_like_count(p) >= 3:
        return False
    if _row_data_likeness(p) >= max(4, (_nonempty_width(p) * 2) // 3):
        return False
    # Current row should look mostly like labels (not percentages / CPT codes grid)
    if _row_data_likeness(q) >= max(4, _nonempty_width(q) - 1):
        return False
    nxt_idx = idx + 1
    if nxt_idx < len(grid):
        # First non-blank row after the header band should look like tabular data.
        nxt_nonempty = idx + 1
        while nxt_nonempty < len(grid):
            cand = _normalize_row_len(grid[nxt_nonempty], width)
            if _nonempty_width(cand) > 0:
                break
            nxt_nonempty += 1
        if nxt_nonempty < len(grid):
            nxt = _normalize_row_len(grid[nxt_nonempty], width)
            if _row_data_likeness(nxt) < max(2, (_nonempty_width(nxt) + 1) // 2):
                return False
    return True


def _maybe_format_percent_cells(hdr: List[str], rows: List[List[Any]]) -> List[List[Any]]:
    """Excel stores percentages as 0.101 for 10.1%; match on-screen conventions for rate/inc columns."""
    if not hdr or not rows:
        return rows
    out: List[List[Any]] = []
    for row in rows:
        nr: List[Any] = []
        for j, val in enumerate(row):
            col = (hdr[j] if j < len(hdr) else "").lower()
            pct_col = (
                "%" in col
                or bool(re.search(r"\brate\b", col))
                or " inc" in col
                or col.endswith(" inc.")
                or col.rstrip(".") == "inc"
            )
            done = False
            if pct_col and isinstance(val, float):
                fv = float(val)
                if fv != fv:  # NaN
                    pass
                elif -0.0001 <= fv <= 1.001:
                    try:
                        if abs(fv) < 1e-12:
                            nr.append("")
                        else:
                            pct = fv * 100.0
                            s_num = ("%g" % round(pct, 6)).rstrip("0").rstrip(".")
                            nr.append(f"{s_num}%")
                        done = True
                    except Exception:
                        done = False
            if not done:
                nr.append(val)
        out.append(nr)
    return out


def _find_likely_grid_header_index(rows: List[List[Any]], *, min_cols: int = 3) -> int:
    """
    Skip narrative / metadata blocks above the first row that behaves like a real grid
    header with data-like rows shortly below (blank spacer rows allowed).
    """
    if not rows:
        return 0
    limit = min(len(rows) - 2, _MAX_SCAN_ROWS)
    best_i = 0
    best_score = -1.0
    for i in range(max(0, limit)):
        r = rows[i]
        ne = _nonempty_width(r)
        if ne < min_cols:
            continue
        max_cell = max((len(str(c).strip()) for c in r), default=0)
        longish = sum(1 for c in r if len(str(c).strip()) > 90)
        if longish >= max(1, ne // 2) and ne <= 4:
            continue
        if ne == 1 and max_cell > 120:
            continue
        pair_sum, top1 = _below_header_data_signal(rows, i)
        hk = _header_keyword_bonus(r)
        dl_here = _row_data_likeness(r)
        rid_here = _row_rate_identifier_count(r)
        num_here = _row_numeric_like_count(r)

        data_signal_ok = pair_sum >= 2 or top1 >= 3
        if hk >= 14.0:
            data_signal_ok = data_signal_ok or top1 >= 2
        if not data_signal_ok:
            continue

        score = ne * 8.0 + min(pair_sum, 18) + hk - (0.3 * longish)
        score -= min(dl_here * 3.8, 28.0)
        score -= min(rid_here * 6.0, 24.0)
        score -= min(max(0, num_here - 1) * 3.0, 18.0)
        if max_cell > 42:
            score -= 1.5
        if score > best_score:
            best_score = score
            best_i = i
    if best_score < 0:
        window = min(80, len(rows))
        cand_rows = [
            i
            for i in range(window)
            if _nonempty_width(rows[i]) >= min_cols
        ]
        if cand_rows:
            mw = max(_nonempty_width(rows[i]) for i in cand_rows)
            tie = [i for i in cand_rows if _nonempty_width(rows[i]) == mw]
            best_i = min(tie, key=lambda j: (_row_data_likeness(rows[j]), -_header_keyword_bonus(rows[j]), j))
        else:
            best_i = 0
    return best_i


def _normalize_row_len(row: List[Any], width: int) -> List[Any]:
    out: List[Any] = []
    for j in range(width):
        if j < len(row):
            v = row[j]
            out.append("" if v is None else v)
        else:
            out.append("")
    return out


def _trim_trailing_empty_columns(rows: List[List[Any]]) -> List[List[Any]]:
    if not rows:
        return rows
    max_j = 0
    for r in rows:
        for j, c in enumerate(r):
            if str(c).strip():
                max_j = max(max_j, j + 1)
    if max_j <= 0:
        return rows
    return [r[:max_j] for r in rows]


def _drop_placeholder_only_columns(hdr: List[str], rows: List[List[Any]]) -> Tuple[List[str], List[List[Any]]]:
    """
    Remove columns where every data cell is blank and the header is synthetic (col_N).
    Also drops columns with entirely blank headers and no body (Excel phantom columns).
    Real headers (e.g. empty 'Notes') are kept so long as they're not col_* placeholders.
    """
    if not hdr and not rows:
        return hdr, rows
    w = max(len(hdr), max((len(r) for r in rows), default=0))
    hdr_pad = [str(hdr[i]) if i < len(hdr) else "" for i in range(w)]
    rows_pad = [_normalize_row_len(r, w) for r in rows]
    keep: List[int] = []
    for j in range(w):
        h = hdr_pad[j].strip()
        has_body = bool(rows_pad and any(str(r[j]).strip() for r in rows_pad))
        if has_body:
            keep.append(j)
            continue
        if _RE_SYNTH_COL_HEADER.match(h):
            continue  # synthetic label, no values — omit
        if not h:
            continue  # no title and no body
        keep.append(j)
    if len(keep) == w:
        return hdr_pad, rows_pad
    new_hdr = [hdr_pad[j] for j in keep]
    new_rows = [[r[j] for j in keep] for r in rows_pad]
    return new_hdr, new_rows


def _finalize_preview_table(hdr: List[Any], rows: List[List[Any]]) -> Tuple[List[str], List[List[Any]]]:
    """Trim trailing empties, then drop col_N columns that have no data."""
    if not hdr and not rows:
        return [], rows
    combined = _trim_trailing_empty_columns([list(hdr)] + list(rows))
    if not combined:
        return [], []
    hdr_s = [str(x) for x in combined[0]]
    body = combined[1:]
    hdr2, body2 = _drop_placeholder_only_columns(hdr_s, body)
    return hdr2, body2


def _table_from_xlsx_merge_aware(
    data: bytes,
    max_rows: int,
    scan_cap: int,
):
    """Full workbook read with merged-cell expansion + two-row header band (fee summary sheets)."""
    try:
        import openpyxl
    except ImportError:
        return None, None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=False, data_only=True)
        try:
            ws = wb.active
            max_r = _ws_scan_last_row(ws, scan_cap)
            max_c = min(int(ws.max_column or 0), 64)
            if max_r < 2 or max_c < 2:
                return None, None

            grid: List[List[Any]] = []
            for r in range(1, max_r + 1):
                row = [ws.cell(r, c).value for c in range(1, max_c + 1)]
                grid.append(row)

            for mrange in list(ws.merged_cells.ranges):
                min_rr, min_cc, max_rr, max_cc = (
                    mrange.min_row,
                    mrange.min_col,
                    mrange.max_row,
                    mrange.max_col,
                )
                v = ws.cell(min_rr, min_cc).value
                for rr in range(min_rr, max_rr + 1):
                    for cc in range(min_cc, max_cc + 1):
                        if 1 <= rr <= max_r and 1 <= cc <= max_c:
                            grid[rr - 1][cc - 1] = v

            for ri, row in enumerate(grid):
                for ci, cell in enumerate(row):
                    if cell is None:
                        grid[ri][ci] = ""

            hdr_idx = _find_likely_grid_header_index(grid)
            width = max(
                len(grid[hdr_idx]),
                max((len(r) for r in grid[hdr_idx:]), default=0),
                max_c,
            )
            width = min(width, max_c)

            combined_headers: Optional[List[str]] = None
            data_start = hdr_idx + 1
            if hdr_idx > 0 and _header_pair_merge_worthy(
                grid[hdr_idx - 1], grid[hdr_idx], width, grid, hdr_idx
            ):
                combined_headers = _combine_two_header_rows(grid[hdr_idx - 1], grid[hdr_idx], width)
                data_start = hdr_idx + 1
            elif hdr_idx > 0 and _nonempty_width(_normalize_row_len(grid[hdr_idx - 1], width)) >= 3:
                dln = (
                    _row_data_likeness(grid[hdr_idx])
                    if hdr_idx < len(grid)
                    else 0
                )
                qln = _nonempty_width(_normalize_row_len(grid[hdr_idx], width))
                if dln <= 2 and qln >= 4:
                    prev_h = [_normalize_cell_header(c) for c in _normalize_row_len(grid[hdr_idx - 1], width)]
                    prev_rn = grid[hdr_idx - 1]
                    prev_ok = (
                        _row_rate_identifier_count(prev_rn) <= 1
                        and _row_numeric_like_count(prev_rn) <= 2
                        and _row_data_likeness(prev_rn) <= max(3, _nonempty_width(prev_rn) // 2)
                    )
                    if prev_ok and any(
                        re.search(r"(workforce|hcbs|rescue|rate increase)", str(x).lower()) for x in prev_h
                    ):
                        combined_headers = _combine_two_header_rows(grid[hdr_idx - 1], grid[hdr_idx], width)
                        data_start = hdr_idx + 1

            if combined_headers:
                hdr_row = combined_headers
            else:
                hdr_row = [_normalize_cell_header(c) for c in _normalize_row_len(grid[hdr_idx], width)]

            hdr_row = _normalize_row_len(hdr_row, width)
            hdr_clean = [
                str(c).strip() if str(c).strip() else f"col_{k + 1}" for k, c in enumerate(hdr_row)
            ]
            rows_out: List[List[Any]] = []
            for r in grid[data_start:]:
                if len(rows_out) >= max_rows:
                    break
                line = _normalize_row_len(r, width)
                if not any(str(c).strip() for c in line):
                    continue
                if str(line[0]).strip().lower().startswith("note"):
                    continue
                rows_out.append(line)

            rows_fmt = _maybe_format_percent_cells(hdr_clean, rows_out)
            nh, nr = _finalize_preview_table(hdr_clean, rows_fmt)
            if not nh:
                return None, None
            return nh, nr
        finally:
            wb.close()
    except Exception:
        return None, None


def _table_from_xlsx(
    data: bytes,
    max_rows: int,
    *,
    max_scan_rows: Optional[int] = None,
):
    try:
        import openpyxl
    except ImportError:
        return None, None

    scan_cap = max_scan_rows if max_scan_rows is not None else _MAX_SCAN_ROWS

    if len(data) <= _RICH_XLSX_BYTES_CAP:
        hit = _table_from_xlsx_merge_aware(data, max_rows, scan_cap)
        if hit[0] is not None:
            return hit

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        raw_scan: List[List[Any]] = []
        for _, row in enumerate(ws.iter_rows(values_only=True)):
            raw_scan.append([(c if c is not None else "") for c in row[:64]])
            if len(raw_scan) >= scan_cap:
                break

        hdr_idx = _find_likely_grid_header_index(raw_scan)
        hdr_row = [_normalize_cell_header(c) for c in raw_scan[hdr_idx]]
        width = max(len(hdr_row), max((len(r) for r in raw_scan[hdr_idx:]), default=0))

        hdr = _normalize_row_len(hdr_row, width)
        hdr = [str(c).strip() if str(c).strip() else f"col_{k + 1}" for k, c in enumerate(hdr)]
        rows_out: List[List[Any]] = []
        for r in raw_scan[hdr_idx + 1 :]:
            if len(rows_out) >= max_rows:
                break
            line = _normalize_row_len(r, width)
            if not any(str(c).strip() for c in line):
                continue
            if str(line[0]).strip().lower().startswith("note"):
                continue
            rows_out.append(line)

        wb.close()
        rows_fmt = _maybe_format_percent_cells(hdr, rows_out)
        nh, nr = _finalize_preview_table(hdr, rows_fmt)
        if not nh:
            return None, None
        return nh, nr
    except Exception:
        return None, None


def _normalize_cell_header(c: Any) -> str:
    if c is None:
        return ""
    return str(c).strip()


def _merge_pdf_vectors(data: bytes) -> List[List[str]]:
    """
    Flatten pdfplumber tables in page order; merge tables matching the widest high-area grid.
    """
    try:
        import pdfplumber
    except ImportError:
        return []

    blocks: List[Tuple[int, List[List[str]]]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables() or []:
                if not tbl:
                    continue
                clean: List[List[str]] = []
                for row in tbl:
                    cells = [(c or "").strip() for c in row]
                    if not any(cells):
                        continue
                    clean.append(cells)
                if len(clean) < 2:
                    continue
                w = max(len(r) for r in clean)
                if w < 3:
                    continue
                area = sum(len(r) for r in clean) * w
                blocks.append((area, clean))

    if not blocks:
        return []

    best_area = max(b[0] for b in blocks)
    dominant_w = max(max(len(r) for r in b[1]) for b in blocks if b[0] == best_area)

    merged: List[List[str]] = []
    tw = dominant_w
    for area, tbl in blocks:
        w_tbl = max(len(r) for r in tbl)
        if area < best_area // 20 and w_tbl < dominant_w - 1:
            continue
        if w_tbl != dominant_w and abs(w_tbl - dominant_w) > 1:
            continue
        for row in tbl:
            padded = row + [""] * tw
            merged.append([str(x) for x in padded[:tw]])
    return merged


def _table_from_pdf_vector(rows: List[List[str]], max_rows: int) -> Tuple[Optional[List[str]], Optional[List[List[Any]]]]:
    if len(rows) < 2:
        return None, None
    hdr_idx = _find_likely_grid_header_index(rows)
    hdr = [_normalize_cell_header(c) for c in rows[hdr_idx]]
    width = max(len(hdr), max(len(r) for r in rows))
    hdr = _normalize_row_len(hdr, width)
    hdr = [str(c).strip() if str(c).strip() else f"col_{k + 1}" for k, c in enumerate(hdr)]
    hk = tuple(str(c).strip().lower() for c in hdr)
    rows_out: List[List[Any]] = []
    for r in rows[hdr_idx + 1 :]:
        if len(rows_out) >= max_rows:
            break
        line = _normalize_row_len(r, width)
        if not any(str(c).strip() for c in line):
            continue
        lk = tuple(str(c).strip().lower() for c in line)
        if lk == hk:
            continue
        rows_out.append(line)
    fh, fb = _finalize_preview_table(hdr, rows_out)
    if not fh:
        return None, None
    return fh, fb


def build_preview_payload(
    resource_url: str,
    *,
    authority: StoredPreviewSession | None,
    referer_override: Optional[str] = None,
    document_hint: Optional[str] = None,
) -> Dict[str, Any]:
    ok, reject = validate_preview_url(resource_url, authority)
    if not ok:
        return {"ok": False, "error_code": reject or "rejected"}

    blob, ctype, status, upstream_hint, fetch_attempts = fetch_resource_bytes(
        resource_url,
        authority,
        referer_override=referer_override,
        document_hint=document_hint,
    )
    if status and status >= 400:
        err = (
            "upstream_auth_xml"
            if upstream_hint == "auth_xml"
            else "upstream_http_status"
        )
        return {
            "ok": False,
            "error_code": err,
            "http_status": status,
            "upstream_attempts": fetch_attempts,
        }
    if not blob:
        return {"ok": False, "error_code": "empty_body", "upstream_attempts": fetch_attempts}

    kind = _kind_from_signals(ctype, resource_url, blob)
    if blob.startswith(b"%PDF"):
        kind = "pdf"
    mime = ctype or ""

    out: Dict[str, Any] = {
        "ok": True,
        "mime": mime,
        "detected_kind": kind,
        "size_bytes": len(blob),
        "warning": None,
    }

    if kind == "probably_html_json":
        snippet = blob[:4000].decode("utf-8", errors="replace")
        out["detected_kind"] = "structured_page"
        out["text_preview"] = snippet
        out["hint"] = "Response looks like HTML/JSON rather than an attachment."
        return out

    if kind == "tabular_text":
        text = blob[:MAX_TEXT_CHARS].decode("utf-8", errors="replace")
        lines = [ln for ln in text.splitlines() if ln.strip()][:420]
        if not lines:
            out["detected_kind"] = "text"
            out["text_preview"] = text[:MAX_TEXT_CHARS]
            return out
        delim = "\t"
        head = lines[0]
        if "\t" not in head and "," in head:
            delim = ","
        elif "; " in head[:220] and "\t" not in head[:220]:
            delim = "; "
        table_rows = [ln.split(delim) for ln in lines]
        mw = max((len(tr) for tr in table_rows), default=0)
        grid = [tr + [""] * (mw - len(tr)) for tr in table_rows]
        hi = _find_likely_grid_header_index(grid, min_cols=2)
        hdr_raw = [_normalize_cell_header(c) for c in grid[hi]]
        hdr = [
            str(h).strip() if str(h).strip() else f"col_{k + 1}" for k, h in enumerate(hdr_raw)
        ]
        body = grid[hi + 1 : hi + 1 + MAX_PREVIEW_ROWS]
        tc, tb = _finalize_preview_table(hdr, body)
        out["detected_kind"] = "delimiter_table"
        out["delimiter"] = delim
        out["table_preview"] = {"columns": tc or [], "rows": tb}
        out["truncated_lines"] = len(lines) >= 400
        out["hint"] = "Leading non-tabular rows were skipped where a main grid could be inferred."
        return out

    if kind == "spreadsheet":
        hdr, rows = _table_from_xlsx(blob, MAX_PREVIEW_ROWS)
        if hdr is not None and rows is not None:
            out["table_preview"] = {"columns": hdr, "rows": rows}
            out["hint"] = "Leading notes/metadata rows were omitted where a grid could be inferred."
            return out
        kind = "binary"

    if kind == "pdf":
        merged = _merge_pdf_vectors(blob)
        ph = pr = None
        if merged:
            ph, pr = _table_from_pdf_vector(merged, MAX_PREVIEW_ROWS)
        if ph is not None and pr is not None:
            out["detected_kind"] = "pdf_tabular"
            out["table_preview"] = {"columns": ph, "rows": pr}
            out["hint"] = (
                "Showing the strongest detected table grid; narrative text above the table is omitted."
            )
            return out
        if len(blob) <= MAX_INLINE_BASE64_RAW:
            out["detected_kind"] = "pdf"
            out["inline_base64"] = base64.standard_b64encode(blob).decode("ascii")
            out["hint"] = (
                "No ruled table grid was extracted from this PDF; showing the rendered file instead."
            )
            return out
        out["detected_kind"] = "binary_large"
        out["mime"] = mime or "application/pdf"
        out["hint"] = (
            "File is too large to render inline here and automated table extraction did not succeed. "
            "Use Proxy download."
        )
        return out

    if kind == "text" or mime.startswith("text/"):
        out["detected_kind"] = "text"
        out["text_preview"] = blob[:MAX_TEXT_CHARS].decode("utf-8", errors="replace")
        return out

    if len(blob) <= MAX_INLINE_BASE64_RAW:
        out["detected_kind"] = "binary_small"
        out["mime"] = mime or "application/octet-stream"
        out["inline_base64"] = base64.standard_b64encode(blob).decode("ascii")

    else:
        out["detected_kind"] = "binary_large"
        out["hint"] = "File exceeds inline preview limit. Use Proxy download."

    return out


def streaming_fetch_resource(
    resource_url: str,
    authority: StoredPreviewSession | None,
    *,
    referer_override: Optional[str] = None,
    document_hint: Optional[str] = None,
):
    ok, reject = validate_preview_url(resource_url, authority)
    if not ok:
        return None, reject

    referer = (referer_override or "").strip()
    if not referer and authority:
        referer = (authority.referrer_url or "").strip()

    session = session_from_browser(authority, extra_referrer=referer_override)
    attempts = accept_attempts_for_preview(resource_url, document_hint=document_hint)
    last_status: Optional[int] = None

    _maybe_warm_referrer(session, resource_url, referer, None)

    for extra_h in attempts:
        merged = _merge_accept_with_browser_headers(resource_url, referer, extra_h)
        resp = session.get(resource_url, stream=True, timeout=120, headers=merged)
        last_status = resp.status_code
        if resp.status_code >= 400:
            resp.close()
            continue

        peek = b""
        try:
            for chunk in resp.iter_content(chunk_size=65536):
                if not chunk:
                    continue
                peek += chunk
                if _xml_body_indicates_auth_failure(peek):
                    resp.close()
                    return None, "upstream_auth_xml"
                if len(peek) >= 8192:
                    break
        except Exception:
            resp.close()
            continue
        if peek and _xml_body_indicates_auth_failure(peek):
            resp.close()
            return None, "upstream_auth_xml"

        def _mk_gen(r):
            def _gen():
                try:
                    yield peek
                    for chunk in r.iter_content(chunk_size=65536):
                        if chunk:
                            yield chunk
                finally:
                    r.close()

            return _gen

        wrapped_resp = resp

        class _PeekedResponse:
            headers = wrapped_resp.headers
            status_code = wrapped_resp.status_code

            def close(self):
                wrapped_resp.close()

            def iter_content(self, chunk_size=65536):
                return _mk_gen(wrapped_resp)()

        return _PeekedResponse(), None

    # curl_cffi buffer fallback (proxy stream)
    ck = _session_cookies_simple_dict(session)
    ck = _curl_warm_referrer_merge_cookies(resource_url, referer, ck, [])
    for caccept in (
        None,
        {
            "Accept": (
                "application/pdf,application/x-pdf,"
                "application/octet-stream;q=0.95,*/*;q=0.05"
            )
        },
    ):
        hit = _download_via_curl_cffi(resource_url, referer, ck, caccept, [])
        if not hit:
            continue
        data, ctype_hdr, _sc = hit
        if data and _xml_body_indicates_auth_failure(data):
            return None, "upstream_auth_xml"

        class _BufferedStream:
            def __init__(self, blob: bytes, ctype_val: str) -> None:
                self._blob = blob
                self.headers = {"content-type": ctype_val}
                self.status_code = 200

            def close(self):
                pass

            def iter_content(self, chunk_size=65536):
                blob = self._blob
                for i in range(0, len(blob), chunk_size):
                    yield blob[i : i + chunk_size]

        return _BufferedStream(data, ctype_hdr or "application/octet-stream"), None

    return None, f"upstream_{last_status or 502}"


def build_artifact_table_preview(
    data: bytes,
    *,
    original_filename: str = "",
    mime_type: str = "",
) -> Dict[str, Any]:
    """
    Parse a saved artifact (PDF / Excel / CSV) into a column + row grid for the SPA fee preview.
    Returns ``{ok, columns, rows}`` or ``{ok: False, error}``.
    """
    name = (original_filename or "").lower()
    mt = (mime_type or "").split(";")[0].strip().lower()

    if not data:
        return {"ok": False, "error": "empty_file"}

    # Companion artifacts are capped at 50 MiB on download; read full buffer for extraction.
    raw = data

    looks_pdf = raw.startswith(b"%PDF") or "pdf" in mt or name.endswith(".pdf")
    if looks_pdf:
        merged = _merge_pdf_vectors(raw)
        if merged:
            ph, pr = _table_from_pdf_vector(merged, ARTIFACT_TABLE_MAX_ROWS)
            if ph is not None and pr is not None:
                return {"ok": True, "columns": ph, "rows": pr}
        return {
            "ok": False,
            "error": "Could not extract a ruled table from this PDF. Use Download to open the file.",
        }

    looks_xlsx = name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in mt or "openxmlformats-officedocument" in mt
    looks_xls = name.endswith(".xls") and not name.endswith(".xlsx")

    if looks_xlsx or _is_ooxlsx_workbook(raw):
        hdr, rows = _table_from_xlsx(
            raw,
            ARTIFACT_TABLE_MAX_ROWS,
            max_scan_rows=min(200_000, ARTIFACT_TABLE_MAX_ROWS + 5000),
        )
        if hdr is not None and rows is not None:
            return {"ok": True, "columns": hdr, "rows": rows}
        return {
            "ok": False,
            "error": "Could not read this spreadsheet (encrypted, corrupt, or missing openpyxl). Use Download.",
        }

    if looks_xls:
        return {"ok": False, "error": "Legacy .xls workbooks cannot be previewed in the app. Use Download."}

    looks_delimited = (
        name.endswith(".csv")
        or "csv" in mt
        or mt.startswith("text/")
        or mt in ("application/csv", "text/csv", "text/tab-separated-values")
    )
    if looks_delimited:
        try:
            text = raw.decode("utf-8-sig", errors="replace")
        except Exception:
            return {"ok": False, "error": "Could not decode file as text."}
        lines = [ln for ln in text.splitlines() if ln.strip()]
        if not lines:
            return {"ok": False, "error": "no_rows"}
        try:
            dia = csv.Sniffer().sniff("\n".join(lines[: min(50, len(lines))]), delimiters=",\t;")
        except csv.Error:
            dia = csv.excel
        rdr = csv.reader(io.StringIO("\n".join(lines)), dia)
        grid: List[List[Any]] = []
        max_grid = ARTIFACT_TABLE_MAX_ROWS + 5000
        for row in rdr:
            if len(grid) >= max_grid:
                break
            grid.append([(c if c is not None else "") for c in row])
        if not grid:
            return {"ok": False, "error": "no_rows"}
        hi = _find_likely_grid_header_index(grid, min_cols=2)
        hdr_raw = [_normalize_cell_header(c) for c in grid[hi]]
        width = max(len(hdr_raw), max((len(r) for r in grid[hi:]), default=0))
        hdr = _normalize_row_len(hdr_raw, width)
        hdr = [str(c).strip() if str(c).strip() else f"col_{k + 1}" for k, c in enumerate(hdr)]
        rows_out: List[List[Any]] = []
        for r in grid[hi + 1 :]:
            if len(rows_out) >= ARTIFACT_TABLE_MAX_ROWS:
                break
            line = _normalize_row_len(r, width)
            if not any(str(c).strip() for c in line):
                continue
            rows_out.append(line)
        tc, tb = _finalize_preview_table(hdr, rows_out)
        if not tc:
            return {"ok": False, "error": "no_table"}
        return {"ok": True, "columns": tc, "rows": tb}

    return {"ok": False, "error": "unsupported_type"}
