"""Write compare diff workbooks to disk (aligned with UI Changed workbook export)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional


def _sheet_name(title: str) -> str:
    t = re.sub(r"[\]\\/*?:\[]", " ", str(title or "sheet")).strip()[:31]
    return t or "Sheet"


def _export_cell(pair_index: int, pair: Dict[str, str], row: Dict[str, Any], *, side: str) -> str:
    fds = row.get("field_diffs")
    if isinstance(fds, list) and pair_index < len(fds):
        fd = fds[pair_index]
        if isinstance(fd, dict):
            key = "state_value" if side == "state" else "dst_value"
            if key in fd:
                return str(fd.get(key) if fd.get(key) is not None else "")
    blob = row.get("state_row" if side == "state" else "dst_row")
    if isinstance(blob, dict):
        col = str(pair.get("state_column" if side == "state" else "dst_column") or "")
        if col in blob:
            return str(blob[col] if blob[col] is not None else "")
        for k, v in blob.items():
            if str(k).strip().lower() == col.lower():
                return str(v if v is not None else "")
    return ""


def _sheet_rows(result: Dict[str, Any], rows: List[Dict[str, Any]], *, side: str) -> List[List[str]]:
    pairs = result.get("column_pairs") or []
    if not isinstance(pairs, list) or not pairs:
        return [["(no mapped columns)"]]
    header = [
        str(p.get("state_column" if side == "state" else "dst_column") or "")
        for p in pairs
        if isinstance(p, dict)
    ]
    lines: List[List[str]] = [header]
    for row in rows:
        if not isinstance(row, dict):
            continue
        lines.append([_export_cell(i, p, row, side=side) for i, p in enumerate(pairs) if isinstance(p, dict)])
    return lines


def export_compare_changes_xlsx(*, compare_result: Dict[str, Any], output_path: Path) -> Optional[Path]:
    """
    Write Modified / Added in State / DST not in State sheets.
    Returns ``output_path`` when written, else ``None`` when there are no diff rows.
    """
    rows = compare_result.get("rows") or []
    if not isinstance(rows, list):
        return None
    mod = [r for r in rows if isinstance(r, dict) and r.get("status") == "mismatch"]
    added = [r for r in rows if isinstance(r, dict) and r.get("status") == "state_only"]
    dst_only = [r for r in rows if isinstance(r, dict) and r.get("status") == "dst_only"]
    if not mod and not added and not dst_only:
        return None

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    specs = [
        ("Modified", _sheet_rows(compare_result, mod, side="state")),
        ("Added in State", _sheet_rows(compare_result, added, side="state")),
        ("DST not in State", _sheet_rows(compare_result, dst_only, side="dst")),
    ]
    for title, aoa in specs:
        if not aoa or len(aoa) <= 1:
            continue
        ws = wb.create_sheet(title=_sheet_name(title))
        for line in aoa:
            ws.append(line)

    if not wb.sheetnames:
        return None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
